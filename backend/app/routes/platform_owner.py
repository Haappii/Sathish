from __future__ import annotations

import json
import re
import secrets
from datetime import date, datetime
from email.message import EmailMessage
from mimetypes import guess_type
from pathlib import Path
from datetime import timedelta
import os
import smtplib

import logging

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from typing import Literal, List, Dict
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import cast, func, text
from sqlalchemy import Date

from app.db import get_db
from app.models.branch import Branch
from app.models.platform_onboard_request import PlatformOnboardRequest
from app.models.platform_payment import PlatformPayment
from app.models.platform_team_profile import PlatformTeamProfile
from app.models.platform_portfolio import PlatformPortfolio
from app.models.platform_user import PlatformUser
from app.models.roles import Role
from app.models.shop_details import ShopDetails
from app.models.support_ticket import SupportTicket
from app.models.system_parameters import SystemParameters
from app.models.users import User
from app.models.subscription_plan import SubscriptionPlan
from app.utils.jwt_token import create_access_token
from app.utils.passwords import encode_password, password_needs_upgrade, verify_password
from app.utils.platform_owner_auth import PlatformOwnerOnly
from app.utils.gcs_upload import upload_file, delete_file, url_to_gcs_path, generate_filename


router = APIRouter(prefix="/platform", tags=["Platform Owner"])

SUPPORT_UPLOADS_DIR = Path("uploads") / "support"
SUPPORT_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
PLATFORM_UPLOADS_DIR = Path("uploads") / "platform"
PLATFORM_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

PORTFOLIO_PARAM_KEY = "portfolio_config"

ABOUT_CONTACT_SETTING_KEYS = {
    "name": "about_contact_name",
    "mobile": "about_contact_mobile",
    "email": "about_contact_email",
    "insta": "about_contact_insta",
    "photo_url": "about_contact_photo_url",
}
PLATFORM_SETTINGS_SHOP_ID = 1

SUPPORT_EMAIL_ENABLED = (os.getenv("SUPPORT_EMAIL_ENABLED") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "y",
}
SENDER_EMAIL = (os.getenv("SUPPORT_SENDER_EMAIL") or "").strip()
SENDER_PASSWORD = (os.getenv("SUPPORT_SENDER_PASSWORD") or "").strip()
SMTP_HOST = (os.getenv("SUPPORT_SMTP_HOST") or "smtp.gmail.com").strip()
SMTP_PORT = int((os.getenv("SUPPORT_SMTP_PORT") or "465").strip())


ALLOWED_BILLING_TYPES = {"store", "hotel"}

# ── Module catalogue ──────────────────────────────────────────────────────────
# These keys match the `key` field in frontend MENU_CATALOG / mobileMenuCatalog.
CORE_MODULES: set[str] = {"sales_billing", "inventory"}   # always on, never configurable

# Modules enabled by default for new shops (platform can still toggle them later).
DEFAULT_ENABLED_MODULES: set[str] = {"billing_history"}

ALL_OPTIONAL_MODULES: list[str] = [
    "cash_drawer", "trends", "analytics", "billing_history",
    "table_billing", "qr_orders", "reservations", "delivery", "recipes",
    "order_live", "kot_management", "online_orders", "advance_orders",
    "offline_sync", "drafts", "returns", "dues", "expenses", "customers",
    "employees", "employee_attendance", "employee_onboarding",
    "loyalty", "gift_cards", "coupons",
    "supplier_ledger", "stock_audit", "item_lots", "labels", "transfers",
    "reports", "feedback_review", "deleted_invoices",
    "alerts", "support_tickets", "admin",
]


def _seed_core_modules(db: Session, shop_id: int) -> None:
    """Seed shop_modules for new shops: core + default-enabled modules on, rest off."""
    all_keys = list(CORE_MODULES) + ALL_OPTIONAL_MODULES
    for key in all_keys:
        enabled = key in CORE_MODULES or key in DEFAULT_ENABLED_MODULES
        db.execute(
            text("""
                INSERT INTO shop_modules (shop_id, module_key, enabled)
                VALUES (:sid, :key, :en)
                ON CONFLICT (shop_id, module_key) DO NOTHING
            """),
            {"sid": shop_id, "key": key, "en": enabled},
        )


def _normalize_billing_type(raw: str | None, *, default: str = "store") -> str:
    """
    Ensure billing/shop type is one of the allowed values.
    Falls back to the provided default when missing.
    """
    val = (raw or "").strip().lower() or (default or "").strip().lower()
    if val == "restaurant":
        val = "hotel"
    if val not in ALLOWED_BILLING_TYPES:
        raise HTTPException(400, "Invalid shop type; choose store or hotel")
    return val


class PlatformLoginIn(BaseModel):
    username: str = Field(min_length=1, max_length=120)
    password: str = Field(min_length=1, max_length=200)


@router.post("/auth/login")
def platform_login(payload: PlatformLoginIn, db: Session = Depends(get_db)):
    username = (payload.username or "").strip()
    password = payload.password or ""
    if not username or not password:
        raise HTTPException(400, "Username and password required")

    # Seed the initial platform owner user on first run.
    any_user = db.query(PlatformUser.platform_user_id).first()
    if any_user is None:
        db.add(
            PlatformUser(
                username="Admin",
                password=encode_password("admin123"),
                status=True,
            )
        )
        db.commit()

    user = (
        db.query(PlatformUser)
        .filter(PlatformUser.username == username, PlatformUser.status == True)
        .first()
    )
    if not user:
        raise HTTPException(403, "Invalid username or password")

    if not verify_password(password, user.password):
        raise HTTPException(403, "Invalid username or password")

    if password_needs_upgrade(user.password):
        user.password = encode_password(password)
        db.commit()

    token = create_access_token(
        {
            "platform_owner": True,
            "platform_username": user.username,
            "platform_user_id": user.platform_user_id,
        }
    )
    return {"access_token": token, "token_type": "bearer"}


class PlatformChangePasswordIn(BaseModel):
    current_password: str = Field(min_length=1, max_length=200)
    new_password: str = Field(min_length=6, max_length=200)


@router.delete("/admin/cleanup-shops")
def cleanup_shops(
    shop_ids: str = Query(..., description="Comma-separated shop IDs"),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    ids = [int(x.strip()) for x in shop_ids.split(",") if x.strip().isdigit()]
    if not ids:
        raise HTTPException(400, "No valid shop IDs")
    if 1 in ids:
        raise HTTPException(400, "Cannot delete shop 1 (primary shop)")

    tables = [
        "invoice_details", "invoice_archive", "invoice_due", "invoice_payment",
        "invoice_discount", "invoice_draft", "invoice",
        "sales_return_items", "sales_return", "sales_return_meta",
        "kot_items", "kot", "order_items", "orders", "table_master",
        "qr_order_items", "qr_orders", "table_qr_sessions", "table_qr_tokens",
        "stock_ledger", "stock_audit", "date_wise_stock", "inventory", "item_lot",
        "stock_transfer_items", "stock_transfers",
        "branch_item_price", "item_prices", "recipe", "modifier", "items", "category",
        "purchase_order_items", "purchase_order_attachments", "purchase_orders",
        "supplier_ledger", "suppliers",
        "advance_order_items", "advance_orders",
        "gift_card_transactions", "gift_cards",
        "coupon_usage", "coupons",
        "loyalty_transactions", "loyalty_accounts",
        "customer_wallet_transactions", "customers",
        "cash_drawer", "branch_expenses", "day_close", "month_close",
        "employee_attendance", "employee_settlements", "employees",
        "online_orders", "delivery", "reservations", "feedback",
        "bulk_import_log", "mail_scheduler", "audit_log", "system_parameters",
        "role_permissions", "support_tickets", "platform_payments",
        "users", "branch", "shop_details",
    ]
    deleted = {}
    for sid in ids:
        deleted[sid] = {}
        for tbl in tables:
            try:
                r = db.execute(text(f"DELETE FROM {tbl} WHERE shop_id = :sid"), {"sid": sid})
                if r.rowcount > 0:
                    deleted[sid][tbl] = r.rowcount
            except Exception:
                db.rollback()
        try:
            r = db.execute(text("DELETE FROM platform_onboard_requests WHERE created_shop_id = :sid"), {"sid": sid})
            if r.rowcount > 0:
                deleted[sid]["platform_onboard_requests"] = r.rowcount
        except Exception:
            db.rollback()
    db.commit()
    return {"success": True, "deleted": deleted}


@router.post("/auth/change-password")
def platform_change_password(
    payload: PlatformChangePasswordIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    user = (
        db.query(PlatformUser)
        .filter(
            PlatformUser.platform_user_id == owner.get("platform_user_id"),
            PlatformUser.status == True,
        )
        .first()
    )
    if not user:
        raise HTTPException(404, "User not found")
    if not verify_password(payload.current_password, user.password):
        raise HTTPException(403, "Current password is incorrect")
    user.password = encode_password(payload.new_password)
    db.commit()
    return {"success": True, "message": "Password updated successfully"}


def _get_param_value(db: Session, shop_id: int, key: str) -> str:
    row = (
        db.query(SystemParameters)
        .filter(SystemParameters.shop_id == shop_id, SystemParameters.param_key == key)
        .first()
    )
    return (row.param_value or "").strip() if row else ""


def _set_param_value(db: Session, shop_id: int, key: str, value: str) -> None:
    row = (
        db.query(SystemParameters)
        .filter(SystemParameters.shop_id == shop_id, SystemParameters.param_key == key)
        .first()
    )
    if row:
        row.param_value = value
    else:
        db.add(SystemParameters(shop_id=shop_id, param_key=key, param_value=value))


def _get_about_contact_payload(db: Session) -> dict:
    return {
        "name": _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["name"]),
        "mobile": _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["mobile"]),
        "email": _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["email"]),
        "insta": _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["insta"]),
        "photo_url": _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["photo_url"]),
    }


@router.get("/public/about-contact")
def public_about_contact(db: Session = Depends(get_db)):
    return _get_about_contact_payload(db)


@router.get("/about-contact")
def get_about_contact(
    name: str | None = Query(None),
    mobile: str | None = Query(None),
    email: str | None = Query(None),
    insta: str | None = Query(None),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    # Legacy-safe fallback: allow updating text fields via GET query params
    # when a proxy/runtime only permits GET for this path.
    if any(v is not None for v in (name, mobile, email, insta)):
        if name is not None:
            _set_param_value(
                db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["name"], (name or "").strip()
            )
        if mobile is not None:
            _set_param_value(
                db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["mobile"], (mobile or "").strip()
            )
        if email is not None:
            _set_param_value(
                db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["email"], (email or "").strip()
            )
        if insta is not None:
            _set_param_value(
                db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["insta"], (insta or "").strip()
            )
        db.commit()
    return _get_about_contact_payload(db)


@router.post("/about-contact")
def update_about_contact(
    name: str = Form(""),
    mobile: str = Form(""),
    email: str = Form(""),
    insta: str = Form(""),
    photo: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    _set_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["name"], (name or "").strip())
    _set_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["mobile"], (mobile or "").strip())
    _set_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["email"], (email or "").strip())
    _set_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["insta"], (insta or "").strip())

    if photo and (photo.filename or "").strip():
        content_type = (photo.content_type or "").lower()
        if not content_type.startswith("image/"):
            raise HTTPException(400, "Photo must be an image")

        ext = Path(photo.filename).suffix.lower() if photo.filename else ""
        if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            ext = ".jpg"

        old_photo_url = _get_param_value(
            db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["photo_url"]
        )
        old_gcs_path = url_to_gcs_path(old_photo_url)
        if old_gcs_path:
            delete_file(old_gcs_path)

        filename = generate_filename("about_contact", ext)
        destination = f"platform/{filename}"
        photo_url = upload_file(photo.file, destination, content_type)
        _set_param_value(db, PLATFORM_SETTINGS_SHOP_ID, ABOUT_CONTACT_SETTING_KEYS["photo_url"], photo_url)

    db.commit()
    return _get_about_contact_payload(db)


@router.put("/about-contact")
def update_about_contact_put(
    name: str = Form(""),
    mobile: str = Form(""),
    email: str = Form(""),
    insta: str = Form(""),
    photo: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    return update_about_contact(
        name=name,
        mobile=mobile,
        email=email,
        insta=insta,
        photo=photo,
        db=db,
        owner=owner,
    )


def _can_send_mail() -> bool:
    # Fall back to sending when SMTP creds exist even if SUPPORT_EMAIL_ENABLED flag is missing.
    return bool(SENDER_EMAIL and SENDER_PASSWORD and SMTP_HOST and SMTP_PORT)


def _send_credentials_email(*, to_email: str, subject: str, content: str, html_content: str | None = None) -> bool:
    if not _can_send_mail():
        return False
    to_addr = (to_email or "").strip()
    if not to_addr or "@" not in to_addr:
        return False

    email = EmailMessage()
    email["Subject"] = subject
    email["From"] = SENDER_EMAIL
    email["To"] = to_addr
    email.set_content(content)
    if html_content:
        email.add_alternative(html_content, subtype="html")

    try:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp.send_message(email)
        return True
    except Exception as e:
        logging.getLogger("uvicorn.error").warning("Email send failed: %s", e)
        return False


def _build_welcome_html(*, shop_name: str, shop_id: int, username: str, password: str, trial_ends: str) -> str:
    return f"""<html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:30px;margin:0;">
  <div style="max-width:520px;margin:auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.1);">
    <div style="background:linear-gradient(135deg,#4338ca,#6366f1,#7c3aed);padding:36px 32px;">
      <div style="font-size:11px;letter-spacing:3px;color:rgba(255,255,255,0.7);text-transform:uppercase;margin-bottom:8px;">Welcome to</div>
      <div style="font-size:28px;font-weight:900;color:#fff;letter-spacing:-0.5px;">Haappii Billing</div>
      <div style="font-size:13px;color:rgba(255,255,255,0.6);margin-top:6px;">Your shop is live and ready to use</div>
    </div>
    <div style="padding:32px;">
      <p style="color:#374151;margin:0 0 20px;font-size:15px;line-height:1.6;">
        Hi! Your shop <strong>{shop_name}</strong> has been activated with a <strong>30-day free trial</strong>. All features are unlocked.
      </p>
      <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:12px;padding:20px;margin-bottom:20px;">
        <table style="width:100%;border-collapse:collapse;font-size:14px;">
          <tr>
            <td style="padding:10px 0;color:#6b7280;border-bottom:1px solid #f3f4f6;">Shop ID</td>
            <td style="padding:10px 0;text-align:right;font-weight:800;color:#111827;border-bottom:1px solid #f3f4f6;font-size:18px;">{shop_id}</td>
          </tr>
          <tr>
            <td style="padding:10px 0;color:#6b7280;border-bottom:1px solid #f3f4f6;">Username</td>
            <td style="padding:10px 0;text-align:right;font-weight:700;color:#111827;border-bottom:1px solid #f3f4f6;font-family:monospace;font-size:15px;">{username}</td>
          </tr>
          <tr>
            <td style="padding:10px 0;color:#6b7280;">Password</td>
            <td style="padding:10px 0;text-align:right;font-weight:700;color:#111827;font-family:monospace;font-size:15px;">{password}</td>
          </tr>
        </table>
      </div>
      <div style="background:linear-gradient(135deg,#ecfdf5,#f0fdf4);border:1.5px solid #a7f3d0;border-radius:10px;padding:14px 18px;text-align:center;margin-bottom:24px;">
        <span style="font-size:13px;font-weight:700;color:#059669;">&#127873; Free trial active until {trial_ends}</span>
      </div>
      <div style="text-align:center;">
        <a href="https://haappiibilling.in/login" style="display:inline-block;background:linear-gradient(135deg,#4338ca,#6366f1);color:#fff;text-decoration:none;padding:14px 36px;border-radius:12px;font-weight:700;font-size:15px;box-shadow:0 4px 16px rgba(99,102,241,0.3);">Login Now &rarr;</a>
      </div>
      <div style="margin-top:24px;padding-top:20px;border-top:1px solid #f3f4f6;">
        <p style="color:#9ca3af;font-size:12px;margin:0 0 6px;">Available on:</p>
        <p style="color:#6b7280;font-size:13px;margin:0;font-weight:600;">&#127760; Web &nbsp;&middot;&nbsp; &#128187; Windows Desktop &nbsp;&middot;&nbsp; &#128241; Android App</p>
      </div>
    </div>
    <div style="background:#f8fafc;padding:18px 32px;text-align:center;border-top:1px solid #e5e7eb;">
      <p style="color:#9ca3af;font-size:11px;margin:0;">This is an automated message from Haappii Billing &middot; <a href="https://haappiibilling.in" style="color:#6366f1;text-decoration:none;">haappiibilling.in</a></p>
    </div>
  </div>
</body></html>"""


class OnboardRequestIn(BaseModel):
    requester_name: str | None = None
    requester_email: str | None = None
    requester_phone: str | None = None
    business: str | None = None
    message: str | None = None

    shop_name: str = Field(min_length=1, max_length=150)
    owner_name: str | None = None
    mobile: str | None = None
    mailid: str | None = None
    gst_number: str | None = None
    billing_type: str | None = "store"
    gst_enabled: bool | None = False
    gst_percent: float | None = 0
    gst_mode: str | None = "inclusive"
    logo_url: str | None = None
    address_line1: str | None = None
    address_line2: str | None = None
    address_line3: str | None = None
    city: str | None = None
    state: str | None = None
    pincode: str | None = None

    branch_name: str = Field(min_length=1, max_length=150)
    branch_address_line1: str | None = None
    branch_address_line2: str | None = None
    branch_city: str | None = None
    branch_state: str | None = None
    branch_country: str | None = None
    branch_pincode: str | None = None

    admin_username: str | None = None
    admin_name: str | None = None


class AcceptOnboardPayload(BaseModel):
    billing_type: str | None = None
    monthly_amount: float | None = None


@router.post("/onboard/requests")
def create_onboard_request(payload: OnboardRequestIn, db: Session = Depends(get_db)):
    """
    Self-service registration: instantly provisions the shop with a 30-day
    free trial and sends login credentials by email.
    """
    from datetime import timedelta

    mobile_digits = "".join(ch for ch in str(payload.mobile or payload.requester_phone or "") if ch.isdigit())
    if len(mobile_digits) < 10:
        raise HTTPException(400, "A valid 10-digit mobile number is required")

    billing_type = _normalize_billing_type(payload.billing_type)
    admin_role = _ensure_admin_role(db)
    _ensure_manager_role(db)
    admin_username = (payload.admin_username or "").strip() or "admin"
    admin_password = _generate_password()

    row = PlatformOnboardRequest(
        status="PENDING",
        created_at=datetime.utcnow(),
        requester_name=(payload.requester_name or "").strip() or None,
        requester_email=(payload.requester_email or "").strip() or None,
        requester_phone=(payload.requester_phone or "").strip() or None,
        business=(payload.business or "").strip() or None,
        message=(payload.message or "").strip() or None,
        shop_name=payload.shop_name.strip(),
        owner_name=(payload.owner_name or "").strip() or None,
        mobile=(payload.mobile or "").strip() or None,
        mailid=(payload.mailid or "").strip() or None,
        gst_number=(payload.gst_number or "").strip() or None,
        billing_type=billing_type,
        gst_enabled=bool(payload.gst_enabled),
        gst_percent=payload.gst_percent or 0,
        gst_mode=(payload.gst_mode or "inclusive"),
        logo_url=(payload.logo_url or "").strip() or None,
        address_line1=payload.address_line1,
        address_line2=payload.address_line2,
        address_line3=payload.address_line3,
        city=payload.city,
        state=payload.state,
        pincode=payload.pincode,
        branch_name=payload.branch_name.strip(),
        branch_address_line1=payload.branch_address_line1,
        branch_address_line2=payload.branch_address_line2,
        branch_city=payload.branch_city,
        branch_state=payload.branch_state,
        branch_country=payload.branch_country,
        branch_pincode=payload.branch_pincode,
        admin_username=admin_username,
        admin_name=(payload.admin_name or "").strip() or None,
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    try:
        shop = ShopDetails(
            shop_name=payload.shop_name.strip(),
            owner_name=(payload.owner_name or "").strip() or None,
            mobile=(payload.mobile or "").strip() or None,
            mailid=(payload.mailid or "").strip() or None,
            billing_type=billing_type,
            gst_enabled=bool(payload.gst_enabled),
            gst_percent=payload.gst_percent or 0,
            gst_mode=(payload.gst_mode or "inclusive"),
            city=payload.city,
            state=payload.state,
            pincode=payload.pincode,
        )
        db.add(shop)
        db.commit()
        db.refresh(shop)

        base_branch_name = payload.branch_name.strip() or "Head Office"
        branch_name = base_branch_name
        existing = db.query(Branch).filter(Branch.branch_name == branch_name).first()
        if existing:
            branch_name = f"{base_branch_name} #{shop.shop_id}"

        branch = Branch(
            shop_id=shop.shop_id,
            branch_name=branch_name,
            city=payload.branch_city,
            state=payload.branch_state,
            country=payload.branch_country,
            pincode=payload.branch_pincode,
            type="Head Office",
            status="ACTIVE",
            branch_close="N",
        )
        db.add(branch)
        db.commit()
        db.refresh(branch)

        trial_end = (datetime.utcnow() + timedelta(days=30)).date()
        shop.head_office_branch_id = branch.branch_id
        shop.expires_on = trial_end
        shop.paid_until = trial_end
        db.add(shop)
        db.commit()
        db.refresh(shop)

        user = User(
            shop_id=shop.shop_id,
            user_name=admin_username,
            password=encode_password(admin_password),
            name=(payload.admin_name or payload.owner_name or admin_username).strip(),
            role=admin_role.role_id,
            status=True,
            branch_id=branch.branch_id,
            must_change_password=True,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        row.status = "ACCEPTED"
        row.decided_at = datetime.utcnow()
        row.decided_by = "auto"
        row.created_shop_id = shop.shop_id
        row.created_branch_id = branch.branch_id
        row.created_admin_user_id = user.user_id
        db.commit()

        recipient = (payload.requester_email or payload.mailid or "").strip()
        email_sent = False
        try:
            trial_display = trial_end.strftime("%d %b %Y")
            plain_content = (
                "Welcome to Haappii Billing!\n\n"
                "Your shop has been activated and is ready to use.\n\n"
                f"Shop ID  : {shop.shop_id}\n"
                f"Username : {admin_username}\n"
                f"Password : {admin_password}\n\n"
                f"Your free trial is active until: {trial_display}\n"
                "All features are unlocked during your trial period.\n\n"
                "Login at: https://haappiibilling.in\n"
            )
            html_content = _build_welcome_html(
                shop_name=payload.shop_name.strip(),
                shop_id=shop.shop_id,
                username=admin_username,
                password=admin_password,
                trial_ends=trial_display,
            )
            email_sent = _send_credentials_email(
                to_email=recipient,
                subject="Your Haappii Billing shop is ready! 🎉",
                content=plain_content,
                html_content=html_content,
            )
        except Exception:
            email_sent = False

        return {
            "success": True,
            "request_id": row.request_id,
            "shop_id": shop.shop_id,
            "username": admin_username,
            "password": admin_password,
            "trial_ends": str(trial_end),
            "email_sent": email_sent,
            "status": "ACTIVATED",
            "message": f"Your shop is live! Login with Shop ID {shop.shop_id}, username '{admin_username}'. Credentials have been sent to {recipient}.",
        }
    except Exception as exc:
        row.status = "PENDING"
        row.decision_note = f"Auto-provision failed: {exc}"
        db.commit()
        return {
            "success": True,
            "request_id": row.request_id,
            "status": "PENDING",
            "message": "Request saved. Our team will activate your shop manually and email credentials shortly.",
        }


@router.get("/onboard/requests")
def list_onboard_requests(
    status: str | None = Query(None),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    q = db.query(PlatformOnboardRequest).order_by(PlatformOnboardRequest.request_id.desc())
    if status:
        q = q.filter(PlatformOnboardRequest.status == status.strip().upper())
    rows = q.limit(int(limit)).all()
    return rows


def _ensure_admin_role(db: Session) -> Role:
    from app.services.role_service import ensure_role

    return ensure_role(db, role_name="Admin")


def _ensure_manager_role(db: Session) -> Role:
    from app.services.role_service import ensure_role

    return ensure_role(db, role_name="Manager")


def _generate_password() -> str:
    # Short enough to type; strong enough for first login.
    return secrets.token_urlsafe(9)


@router.post("/onboard/requests/{request_id}/accept")
def accept_onboard_request(
    request_id: int,
    payload: AcceptOnboardPayload | None = None,
    note: str | None = None,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(PlatformOnboardRequest).filter(PlatformOnboardRequest.request_id == request_id).first()
    if not row:
        raise HTTPException(404, "Request not found")
    if row.status != "PENDING":
        raise HTTPException(400, f"Request already {row.status}")

    confirmed_billing_type = _normalize_billing_type(
        payload.billing_type if payload else row.billing_type,
        default=row.billing_type or "store",
    )
    row.billing_type = confirmed_billing_type
    monthly_amount = None
    if payload and payload.monthly_amount is not None:
        try:
            monthly_amount = round(float(payload.monthly_amount), 2)
        except Exception:
            raise HTTPException(400, "Invalid monthly amount")
    row.decision_note = (note or "").strip() or None
    if monthly_amount is not None:
        extra = f"Monthly Amount: {monthly_amount}"
        row.decision_note = f"{row.decision_note + ' | ' if row.decision_note else ''}{extra}"

    admin_role = _ensure_admin_role(db)
    _ensure_manager_role(db)
    admin_username = (row.admin_username or "admin").strip()
    admin_password = _generate_password()

    try:
        shop = ShopDetails(
            shop_name=(row.shop_name or "").strip(),
            owner_name=row.owner_name,
            mobile=row.mobile,
            mailid=row.mailid,
            gst_number=row.gst_number,
            billing_type=confirmed_billing_type,
            gst_enabled=bool(row.gst_enabled),
            gst_percent=row.gst_percent or 0,
            gst_mode=row.gst_mode or "inclusive",
            logo_url=row.logo_url,
            address_line1=row.address_line1,
            address_line2=row.address_line2,
            address_line3=row.address_line3,
            city=row.city,
            state=row.state,
            pincode=row.pincode,
        )
        db.add(shop)
        db.commit()
        db.refresh(shop)

        # Ensure branch name uniqueness (DB has uq on branch_name)
        base_branch_name = (row.branch_name or "").strip() or "Head Office"
        branch_name = base_branch_name
        existing = db.query(Branch).filter(Branch.branch_name == branch_name).first()
        if existing:
            branch_name = f"{base_branch_name} #{shop.shop_id}"

        branch = Branch(
            shop_id=shop.shop_id,
            branch_name=branch_name,
            address_line1=row.branch_address_line1,
            address_line2=row.branch_address_line2,
            city=row.branch_city,
            state=row.branch_state,
            country=row.branch_country,
            pincode=row.branch_pincode,
            type="Head Office",
            status="ACTIVE",
            branch_close="N",
        )
        db.add(branch)
        db.commit()
        db.refresh(branch)

        shop.head_office_branch_id = branch.branch_id
        from datetime import timedelta
        trial_end = (datetime.utcnow() + timedelta(days=30)).date()
        shop.expires_on = trial_end
        shop.paid_until = trial_end
        db.add(shop)
        db.commit()
        db.refresh(shop)

        exists = db.query(User).filter(User.shop_id == shop.shop_id, User.user_name == admin_username).first()
        if exists:
            raise HTTPException(400, "Admin username already exists")

        user = User(
            shop_id=shop.shop_id,
            user_name=admin_username,
            password=encode_password(admin_password),
            name=row.admin_name or admin_username,
            role=admin_role.role_id,
            status=True,
            branch_id=branch.branch_id,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        row.status = "ACCEPTED"
        row.decided_at = datetime.utcnow()
        row.decided_by = str(owner.get("platform_username") or "")
        row.created_shop_id = shop.shop_id
        row.created_branch_id = branch.branch_id
        row.created_admin_user_id = user.user_id
        db.commit()
        db.refresh(row)

        recipient = (row.requester_email or row.mailid or "").strip()
        email_sent = False
        try:
            content = (
                "Welcome to Haappii Billing!\n\n"
                "Your shop has been activated and is ready to use.\n\n"
                f"Shop ID  : {shop.shop_id}\n"
                f"Username : {admin_username}\n"
                f"Password : {admin_password}\n\n"
            )
            if monthly_amount is not None:
                content += f"Monthly Amount: ₹{monthly_amount}\n\n"
            content += (
                f"Your free trial is active until: {trial_end.strftime('%d %b %Y')}\n"
                "All features are unlocked during your trial period.\n\n"
                "Login at: https://haappiibilling.in\n"
            )

            email_sent = _send_credentials_email(
                to_email=recipient,
                subject="Your Haappii Billing shop is ready!",
                content=content,
            )
        except Exception:
            # Don't fail provisioning due to email issues.
            email_sent = False

        return {
            "success": True,
            "request_id": row.request_id,
            "status": row.status,
            "shop_id": shop.shop_id,
            "branch_id": branch.branch_id,
            "admin_username": admin_username,
            "admin_password": admin_password,
            "email_sent": email_sent,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Setup failed: {str(e)}")


@router.post("/onboard/requests/{request_id}/reject")
def reject_onboard_request(
    request_id: int,
    note: str | None = None,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(PlatformOnboardRequest).filter(PlatformOnboardRequest.request_id == request_id).first()
    if not row:
        raise HTTPException(404, "Request not found")
    if row.status != "PENDING":
        raise HTTPException(400, f"Request already {row.status}")

    row.status = "REJECTED"
    row.decided_at = datetime.utcnow()
    row.decided_by = str(owner.get("platform_username") or "")
    row.decision_note = (note or "").strip() or None
    db.commit()
    db.refresh(row)

    recipient = (row.requester_email or row.mailid or "").strip()
    if recipient:
        try:
            content = (
                "We are unable to approve your onboarding request at this time.\n\n"
                f"Request ID: {row.request_id}\n"
            )
            if row.decision_note:
                content += f"Reason: {row.decision_note}\n\n"
            content += "You may reply to this email for clarification or submit a new request."

            _send_credentials_email(
                to_email=recipient,
                subject="Your onboarding request was rejected",
                content=content,
            )
        except Exception:
            pass
    return {"success": True, "request_id": row.request_id, "status": row.status}


@router.get("/support/tickets")
def platform_list_tickets(
    ticket_type: str | None = None,
    status: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    q = db.query(SupportTicket).order_by(SupportTicket.ticket_id.desc())
    if ticket_type:
        q = q.filter(SupportTicket.ticket_type == ticket_type.upper())
    if status:
        q = q.filter(SupportTicket.status == status.upper())
    return q.limit(min(max(int(limit), 1), 500)).all()


@router.get("/revenue")
def platform_revenue(
    days: int = Query(30, ge=1, le=3650),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    since = datetime.utcnow() - timedelta(days=int(days))
    amt = (
        db.query(func.coalesce(func.sum(PlatformPayment.amount), 0))
        .filter(PlatformPayment.paid_on >= since)
        .scalar()
        or 0
    )
    return {"days": int(days), "total": float(amt)}


@router.get("/revenue/daily")
def platform_revenue_daily(
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    since = datetime.utcnow() - timedelta(days=int(days))
    rows = (
        db.query(
            cast(func.date(PlatformPayment.paid_on), Date).label("day"),
            func.coalesce(func.sum(PlatformPayment.amount), 0).label("revenue"),
        )
        .filter(PlatformPayment.paid_on >= since)
        .group_by(cast(func.date(PlatformPayment.paid_on), Date))
        .order_by(cast(func.date(PlatformPayment.paid_on), Date))
        .all()
    )
    return [{"date": str(r.day), "revenue": float(r.revenue or 0)} for r in rows]


@router.get("/plans")
def list_plans(
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    q = db.query(SubscriptionPlan)
    if not include_inactive:
        q = q.filter(SubscriptionPlan.is_active == True)  # noqa: E712
    rows = q.order_by(SubscriptionPlan.plan_id.desc()).all()
    return [
        {
            "plan_id": r.plan_id,
            "name": r.name,
            "duration_months": r.duration_months,
            "price": float(r.price or 0),
            "is_active": bool(r.is_active),
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/plans")
def create_plan(payload: PlanCreateIn, db: Session = Depends(get_db), owner=Depends(PlatformOwnerOnly)):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(400, "Name required")
    exists = db.query(SubscriptionPlan).filter(SubscriptionPlan.name.ilike(name)).first()
    if exists:
        raise HTTPException(400, "Plan name already exists")
    row = SubscriptionPlan(
        name=name,
        duration_months=int(payload.duration_months),
        price=float(payload.price),
        is_active=True,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"success": True, "plan_id": row.plan_id}


@router.post("/plans/{plan_id}/status")
def update_plan_status(
    plan_id: int,
    is_active: bool = True,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(SubscriptionPlan).filter(SubscriptionPlan.plan_id == plan_id).first()
    if not row:
        raise HTTPException(404, "Plan not found")
    row.is_active = bool(is_active)
    db.commit()
    db.refresh(row)
    return {"success": True, "plan_id": row.plan_id, "is_active": row.is_active}


@router.get("/shops/{shop_id}/detail")
def platform_shop_detail(
    shop_id: int,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")
    today = datetime.utcnow().date()
    expires = getattr(s, "expires_on", None)
    paid_until = getattr(s, "paid_until", None)
    plan_upper = (getattr(s, "plan", None) or "TRIAL").upper()
    status = "ACTIVE"
    if plan_upper == "DISABLED":
        status = "DISABLED"
    elif expires and today > expires:
        status = "EXPIRED"
    elif paid_until and today > paid_until:
        status = "EXPIRED"
    elif plan_upper == "TRIAL":
        status = "TRIAL"
    return {
        "shop_id": s.shop_id,
        "shop_name": s.shop_name,
        "owner_name": s.owner_name,
        "mobile": s.mobile,
        "mailid": s.mailid,
        "billing_type": s.billing_type,
        "address_line1": s.address_line1,
        "address_line2": s.address_line2,
        "address_line3": s.address_line3,
        "city": s.city,
        "state": s.state,
        "pincode": s.pincode,
        "plan": plan_upper,
        "expires_on": expires,
        "paid_until": paid_until,
        "last_payment_on": getattr(s, "last_payment_on", None),
        "total_paid": float(getattr(s, "total_paid", 0) or 0),
        "status": status,
        "max_branches": getattr(s, "max_branches", None),
        "max_users": getattr(s, "max_users", None),
    }


@router.post("/shops/{shop_id}/billing-type")
def update_shop_billing_type(
    shop_id: int,
    payload: BillingTypeIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")
    new_type = _normalize_billing_type(payload.billing_type)
    s.billing_type = new_type
    db.commit()
    db.refresh(s)
    return {"success": True, "billing_type": s.billing_type}


@router.get("/shops")
def platform_list_shops(
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    today = datetime.utcnow().date()
    shops = db.query(ShopDetails).order_by(ShopDetails.shop_id.desc()).all()
    out = []
    for s in shops:
        expires = getattr(s, "expires_on", None)
        paid_until = getattr(s, "paid_until", None)
        status = "ACTIVE"
        plan_upper = (getattr(s, "plan", None) or "TRIAL").upper()
        if plan_upper == "DISABLED":
            status = "DISABLED"
        elif expires and today > expires:
            status = "EXPIRED"
        elif paid_until and today > paid_until:
            status = "EXPIRED"
        elif plan_upper == "TRIAL":
            status = "TRIAL"

        out.append(
            {
                "shop_id": s.shop_id,
                "shop_name": s.shop_name,
            "mailid": getattr(s, "mailid", None),
            "mobile": getattr(s, "mobile", None),
            "billing_type": getattr(s, "billing_type", None),
            "is_demo": bool(getattr(s, "is_demo", False)),
            "expires_on": str(expires) if expires else None,
            "plan": plan_upper or "TRIAL",
                "last_payment_on": str(getattr(s, "last_payment_on", None)) if getattr(s, "last_payment_on", None) else None,
                "next_renewal": str(paid_until) if paid_until else None,
                "total_paid": float(getattr(s, "total_paid", 0) or 0),
                "status": status,
                "revenue": float(getattr(s, "total_paid", 0) or 0),
            }
        )
    return out


class ShopPaymentUpdateIn(BaseModel):
    plan: str | None = None
    plan_id: int | None = None
    extend_days: int | None = Field(default=None, ge=1, le=3650)
    paid_until: str | None = None  # YYYY-MM-DD
    amount: float | None = Field(default=None, ge=0)


class ShopLimitsUpdateIn(BaseModel):
    max_branches: int | None = Field(default=None, ge=1, le=500)
    max_users: int | None = Field(default=None, ge=1, le=500)


@router.api_route("/shops/{shop_id}/update-limits", methods=["POST", "PUT"])
@router.api_route("/shops/{shop_id}/update-limits/", methods=["POST", "PUT"], include_in_schema=False)
def update_shop_limits(
    shop_id: int,
    payload: ShopLimitsUpdateIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")
    # Allow None to clear the limit (unlimited)
    s.max_branches = payload.max_branches
    s.max_users = payload.max_users
    db.commit()
    db.refresh(s)
    return {
        "success": True,
        "max_branches": s.max_branches,
        "max_users": s.max_users,
    }


class ShopStatusUpdateIn(BaseModel):
    status: Literal["ACTIVE", "DISABLED"]


class PlanCreateIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    duration_months: int = Field(ge=1, le=36)
    price: float = Field(ge=0)


class BillingTypeIn(BaseModel):
    billing_type: str


@router.post("/shops/{shop_id}/update-payment")
def update_shop_payment(
    shop_id: int,
    payload: ShopPaymentUpdateIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")

    today = datetime.utcnow().date()
    # If plan_id is provided, it takes precedence and auto-sets plan, paid_until, total_paid, last_payment_on
    if payload.plan_id is not None:
        plan_row = (
            db.query(SubscriptionPlan)
            .filter(SubscriptionPlan.plan_id == int(payload.plan_id), SubscriptionPlan.is_active == True)  # noqa: E712
            .first()
        )
        if not plan_row:
            raise HTTPException(404, "Plan not found or inactive")
        plan_name = (plan_row.name or "").strip().upper()[:30] or "PLAN"
        plan_price = float(plan_row.price or 0)
        s.plan = plan_name
        days = int(plan_row.duration_months) * 30
        s.paid_until = today + timedelta(days=days)
        s.total_paid = float(getattr(s, "total_paid", 0) or 0) + plan_price
        s.last_payment_on = today
        if plan_price > 0:
            db.add(PlatformPayment(
                shop_id=shop_id, amount=plan_price,
                plan_name=plan_name, note=f"Plan: {plan_name} ({plan_row.duration_months}mo)",
            ))
    elif payload.plan:
        s.plan = (payload.plan or "").strip().upper()[:30] or s.plan

    if payload.paid_until:
        try:
            y, m, d = [int(x) for x in str(payload.paid_until).split("-")]
            s.paid_until = date(y, m, d)
        except Exception:
            raise HTTPException(400, "paid_until must be YYYY-MM-DD")
    elif payload.extend_days:
        base = s.paid_until if getattr(s, "paid_until", None) and s.paid_until > today else today
        s.paid_until = base + timedelta(days=int(payload.extend_days))

    if payload.amount is not None and float(payload.amount or 0) > 0:
        pay_amt = float(payload.amount)
        s.total_paid = float(getattr(s, "total_paid", 0) or 0) + pay_amt
        s.last_payment_on = today
        db.add(PlatformPayment(
            shop_id=shop_id, amount=pay_amt,
            plan_name=s.plan, note="Manual payment",
        ))

    db.commit()
    db.refresh(s)
    return {"success": True}


@router.post("/shops/{shop_id}/status")
def update_shop_status(
    shop_id: int,
    payload: ShopStatusUpdateIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")

    status = (payload.status or "").strip().upper()
    today = datetime.utcnow().date()
    if status == "DISABLED":
        s.plan = "DISABLED"
        s.expires_on = today - timedelta(days=1)
    elif status == "TRIAL":
        s.plan = "TRIAL"
        s.expires_on = today + timedelta(days=30)
    else:
        s.plan = "ACTIVE"
        s.expires_on = None

    db.commit()
    db.refresh(s)
    return {"success": True, "status": status}


@router.post("/shops/{shop_id}/reminder")
def send_shop_reminder(
    shop_id: int,
    kind: str = Query("PAYMENT_DUE"),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    s = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not s:
        raise HTTPException(404, "Shop not found")

    to_email = (getattr(s, "mailid", None) or "").strip()
    paid_until = getattr(s, "paid_until", None)
    subject = "Payment reminder"
    content = (
        "Reminder from the platform.\n\n"
        f"Shop: {getattr(s, 'shop_name', '')}\n"
        f"Shop ID: {s.shop_id}\n"
        f"Plan: {getattr(s, 'plan', 'TRIAL')}\n"
        + (f"Paid until: {paid_until}\n" if paid_until else "")
        + "\nPlease renew your subscription to avoid access interruption.\n"
    )
    email_sent = False
    try:
        email_sent = _send_credentials_email(to_email=to_email, subject=subject, content=content)
    except Exception:
        email_sent = False

    return {"success": True, "email_sent": bool(email_sent)}


@router.post("/support/tickets/{ticket_id}/status")
def platform_update_ticket_status(
    ticket_id: int,
    new_status: str,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(SupportTicket).filter(SupportTicket.ticket_id == ticket_id).first()
    if not row:
        raise HTTPException(404, "Ticket not found")

    row.status = (new_status or "").strip().upper() or row.status
    db.commit()
    db.refresh(row)
    return {"success": True, "ticket_id": row.ticket_id, "status": row.status}


def _demo_expiry(days: int) -> date:
    d = int(days)
    if d < 1 or d > 365:
        raise HTTPException(400, "Expiry days must be between 1 and 365")
    return (datetime.utcnow().date() + timedelta(days=d))


@router.post("/demo/tickets/{ticket_id}/accept")
def accept_demo_ticket(
    ticket_id: int,
    days: int = Query(7, ge=1, le=365),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    t = db.query(SupportTicket).filter(SupportTicket.ticket_id == ticket_id).first()
    if not t:
        raise HTTPException(404, "Ticket not found")
    if str(t.ticket_type or "").upper() != "DEMO":
        raise HTTPException(400, "Not a demo ticket")
    if str(t.status or "").upper() != "OPEN":
        raise HTTPException(400, f"Ticket already {t.status}")

    expires_on = _demo_expiry(days)

    # Use business / name as shop name fallback.
    shop_name = (t.business or "").strip() or (t.user_name or "").strip() or f"Demo Shop {ticket_id}"
    admin_username = "admin"
    admin_password = _generate_password()

    admin_role = _ensure_admin_role(db)
    _ensure_manager_role(db)

    try:
        shop = ShopDetails(
            shop_name=shop_name,
            owner_name=t.user_name,
            mobile=getattr(t, "phone", None) or None,
            mailid=getattr(t, "email", None) or None,
            billing_type="store",
            gst_enabled=False,
            gst_percent=0,
            gst_mode="inclusive",
            is_demo=True,
            expires_on=expires_on,
        )
        db.add(shop)
        db.commit()
        db.refresh(shop)

        branch = Branch(
            shop_id=shop.shop_id,
            branch_name="Head Office",
            type="Head Office",
            status="ACTIVE",
            branch_close="N",
        )
        db.add(branch)
        db.commit()
        db.refresh(branch)

        shop.head_office_branch_id = branch.branch_id
        db.add(shop)
        db.commit()
        db.refresh(shop)

        user = User(
            shop_id=shop.shop_id,
            user_name=admin_username,
            password=encode_password(admin_password),
            name="Demo Admin",
            role=admin_role.role_id,
            status=True,
            branch_id=branch.branch_id,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        t.status = "RESOLVED"
        t.provisioned_shop_id = shop.shop_id
        t.provisioned_branch_id = branch.branch_id
        t.provisioned_admin_user_id = user.user_id
        t.provisioned_expires_on = expires_on
        t.decided_by = str(owner.get("platform_username") or "")
        t.decided_at = datetime.utcnow()
        db.commit()

        email_sent = False
        try:
            email_sent = _send_credentials_email(
                to_email=(getattr(t, "email", None) or ""),
                subject="Your demo is activated",
                content=(
                    "Your demo request is approved.\n\n"
                    f"Shop ID: {shop.shop_id}\n"
                    f"Username: {admin_username}\n"
                    f"Password: {admin_password}\n"
                    f"Expires on: {expires_on}\n\n"
                    "Login URL: / (open the app and login)\n"
                ),
            )
        except Exception:
            email_sent = False

        return {
            "success": True,
            "ticket_id": t.ticket_id,
            "shop_id": shop.shop_id,
            "branch_id": branch.branch_id,
            "admin_username": admin_username,
            "admin_password": admin_password,
            "expires_on": str(expires_on),
            "email_sent": email_sent,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Failed to provision demo: {str(e)}")


@router.post("/demo/tickets/{ticket_id}/reject")
def reject_demo_ticket(
    ticket_id: int,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    t = db.query(SupportTicket).filter(SupportTicket.ticket_id == ticket_id).first()
    if not t:
        raise HTTPException(404, "Ticket not found")
    if str(t.ticket_type or "").upper() != "DEMO":
        raise HTTPException(400, "Not a demo ticket")
    if str(t.status or "").upper() != "OPEN":
        raise HTTPException(400, f"Ticket already {t.status}")

    t.status = "CLOSED"
    t.decided_by = str(owner.get("platform_username") or "")
    t.decided_at = datetime.utcnow()
    db.commit()
    return {"success": True, "ticket_id": t.ticket_id, "status": t.status}


@router.get("/support/tickets/{ticket_id}/attachment")
def platform_download_ticket_attachment(
    ticket_id: int,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(SupportTicket).filter(SupportTicket.ticket_id == ticket_id).first()
    if not row or not row.attachment_path:
        raise HTTPException(404, "Attachment not found")

    p = Path(row.attachment_path)
    try:
        resolved = p.resolve()
        resolved.relative_to(SUPPORT_UPLOADS_DIR.resolve())
    except Exception:
        raise HTTPException(400, "Invalid attachment path")

    if not resolved.exists():
        raise HTTPException(404, "Attachment file missing")

    mime_type, _ = guess_type(str(resolved))
    return FileResponse(
        str(resolved),
        filename=(row.attachment_filename or resolved.name),
        media_type=(mime_type or "application/octet-stream"),
    )


# ── Shop Modules management ───────────────────────────────────────────────────

class ShopModulesIn(BaseModel):
    modules: Dict[str, bool]  # {module_key: enabled}


@router.get("/shops/{shop_id}/modules")
def get_shop_module_config(
    shop_id: int,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    """Return current module enable/disable state for a shop."""
    if not db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first():
        raise HTTPException(404, "Shop not found")

    rows = db.execute(
        text("SELECT module_key, enabled FROM shop_modules WHERE shop_id = :sid"),
        {"sid": shop_id},
    ).fetchall()

    if not rows:
        # Not configured yet → all optional modules default ON (backward compat)
        modules = {m: True for m in list(CORE_MODULES) + ALL_OPTIONAL_MODULES}
        return {"shop_id": shop_id, "configured": False, "modules": modules}

    saved = {r.module_key: r.enabled for r in rows}
    modules: Dict[str, bool] = {}
    for m in CORE_MODULES:
        modules[m] = True                          # core always on
    for m in ALL_OPTIONAL_MODULES:
        modules[m] = saved.get(m, False)           # default OFF if not seeded

    return {"shop_id": shop_id, "configured": True, "modules": modules}


@router.post("/shops/{shop_id}/modules")
def save_shop_module_config(
    shop_id: int,
    payload: ShopModulesIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    """Enable or disable feature modules for a shop."""
    if not db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first():
        raise HTTPException(404, "Shop not found")

    for key, enabled in (payload.modules or {}).items():
        if key in CORE_MODULES:
            continue   # core modules cannot be disabled
        if key not in ALL_OPTIONAL_MODULES:
            continue   # ignore unknown keys
        db.execute(
            text("""
                INSERT INTO shop_modules (shop_id, module_key, enabled)
                VALUES (:sid, :key, :en)
                ON CONFLICT (shop_id, module_key) DO UPDATE SET enabled = EXCLUDED.enabled
            """),
            {"sid": shop_id, "key": key, "en": bool(enabled)},
        )
    db.commit()
    return {"success": True}


# ── Direct shop creation (no onboard request needed) ─────────────────────────

class DirectCreateShopIn(BaseModel):
    shop_name: str = Field(min_length=1, max_length=150)
    owner_name: str | None = None
    mobile: str | None = None
    mailid: str | None = None
    billing_type: str = "store"
    branch_name: str = "Head Office"
    admin_username: str = "admin"
    admin_name: str | None = None
    gst_enabled: bool = False
    gst_percent: float = 0
    gst_mode: str = "inclusive"
    address_line1: str | None = None
    city: str | None = None
    state: str | None = None
    pincode: str | None = None


@router.post("/shops/create")
def direct_create_shop(
    payload: DirectCreateShopIn,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    """
    Create a new shop instantly (no onboarding approval flow).
    Credentials are emailed to mailid if SMTP is configured.
    The new shop starts with only core modules (sales_billing + inventory) enabled.
    """
    confirmed_billing_type = _normalize_billing_type(payload.billing_type)
    admin_role = _ensure_admin_role(db)
    _ensure_manager_role(db)
    admin_username = (payload.admin_username or "admin").strip()
    admin_password = _generate_password()

    try:
        shop = ShopDetails(
            shop_name=payload.shop_name.strip(),
            owner_name=(payload.owner_name or "").strip() or None,
            mobile=(payload.mobile or "").strip() or None,
            mailid=(payload.mailid or "").strip() or None,
            billing_type=confirmed_billing_type,
            gst_enabled=bool(payload.gst_enabled),
            gst_percent=payload.gst_percent or 0,
            gst_mode=payload.gst_mode or "inclusive",
            address_line1=payload.address_line1,
            city=payload.city,
            state=payload.state,
            pincode=payload.pincode,
        )
        db.add(shop)
        db.commit()
        db.refresh(shop)

        base_branch_name = (payload.branch_name or "Head Office").strip()
        branch_name = base_branch_name
        if db.query(Branch).filter(Branch.branch_name == branch_name).first():
            branch_name = f"{base_branch_name} #{shop.shop_id}"

        branch = Branch(
            shop_id=shop.shop_id,
            branch_name=branch_name,
            type="Head Office",
            status="ACTIVE",
            branch_close="N",
        )
        db.add(branch)
        db.commit()
        db.refresh(branch)

        shop.head_office_branch_id = branch.branch_id
        db.commit()

        if db.query(User).filter(User.shop_id == shop.shop_id, User.user_name == admin_username).first():
            raise HTTPException(400, "Admin username already exists for this shop")

        user = User(
            shop_id=shop.shop_id,
            user_name=admin_username,
            password=encode_password(admin_password),
            name=payload.admin_name or admin_username,
            role=admin_role.role_id,
            status=True,
            branch_id=branch.branch_id,
        )
        db.add(user)
        db.commit()

        recipient = (payload.mailid or "").strip()
        email_sent = False
        try:
            content = (
                "Your shop has been created on Haappii Billing.\n\n"
                f"Shop: {shop.shop_name}\n"
                f"Shop ID: {shop.shop_id}\n"
                f"Username: {admin_username}\n"
                f"Password: {admin_password}\n\n"
                "Login and start billing right away!\n"
            )
            email_sent = _send_credentials_email(
                to_email=recipient,
                subject=f"Welcome to Haappii Billing – {shop.shop_name}",
                content=content,
            )
        except Exception:
            email_sent = False

        return {
            "success": True,
            "shop_id": shop.shop_id,
            "branch_id": branch.branch_id,
            "admin_username": admin_username,
            "admin_password": admin_password,
            "email_sent": email_sent,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Shop creation failed: {str(e)}")


# ── Team Profiles ─────────────────────────────────────────────────────────────

TEAM_UPLOADS_DIR = Path("uploads") / "team"
TEAM_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def _profile_row(p: PlatformTeamProfile, portfolio_slug: str | None = None) -> dict:
    return {
        "profile_id": p.profile_id,
        "name": p.name,
        "role_title": p.role_title or "",
        "bio": p.bio or "",
        "photo_url": p.photo_url or "",
        "display_order": p.display_order,
        "is_active": p.is_active,
        "created_at": str(p.created_at) if p.created_at else None,
        "portfolio_slug": portfolio_slug,
    }


def _portfolio_slug_map(db: Session, profile_ids: list[int], active_only: bool) -> dict[int, str]:
    if not profile_ids:
        return {}
    q = db.query(PlatformPortfolio).filter(PlatformPortfolio.profile_id.in_(profile_ids))
    if active_only:
        q = q.filter(PlatformPortfolio.is_active == True)  # noqa: E712
    return {row.profile_id: row.slug for row in q.all() if row.profile_id is not None}


@router.get("/public/team-profiles")
def public_team_profiles(db: Session = Depends(get_db)):
    rows = (
        db.query(PlatformTeamProfile)
        .filter(PlatformTeamProfile.is_active == True)  # noqa: E712
        .order_by(PlatformTeamProfile.display_order.asc(), PlatformTeamProfile.profile_id.asc())
        .all()
    )
    slug_map = _portfolio_slug_map(db, [p.profile_id for p in rows], active_only=True)
    return [_profile_row(p, slug_map.get(p.profile_id)) for p in rows]


@router.get("/team-profiles")
def list_team_profiles(
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    rows = (
        db.query(PlatformTeamProfile)
        .order_by(PlatformTeamProfile.display_order.asc(), PlatformTeamProfile.profile_id.asc())
        .all()
    )
    slug_map = _portfolio_slug_map(db, [p.profile_id for p in rows], active_only=False)
    return [_profile_row(p, slug_map.get(p.profile_id)) for p in rows]


@router.post("/team-profiles")
def create_team_profile(
    name: str = Form(...),
    role_title: str = Form(""),
    bio: str = Form(""),
    display_order: int = Form(0),
    is_active: bool = Form(True),
    photo: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    name = (name or "").strip()
    if not name:
        raise HTTPException(400, "Name is required")

    photo_url = None
    if photo and (photo.filename or "").strip():
        content_type = (photo.content_type or "").lower()
        if not content_type.startswith("image/"):
            raise HTTPException(400, "Photo must be an image")
        ext = Path(photo.filename).suffix.lower() if photo.filename else ".jpg"
        if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            ext = ".jpg"
        filename = generate_filename("team", ext)
        destination = f"team/{filename}"
        photo_url = upload_file(photo.file, destination, content_type)

    row = PlatformTeamProfile(
        name=name,
        role_title=(role_title or "").strip(),
        bio=(bio or "").strip() or None,
        photo_url=photo_url,
        display_order=int(display_order),
        is_active=bool(is_active),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _profile_row(row)


@router.put("/team-profiles/{profile_id}")
def update_team_profile(
    profile_id: int,
    name: str = Form(...),
    role_title: str = Form(""),
    bio: str = Form(""),
    display_order: int = Form(0),
    is_active: bool = Form(True),
    photo: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == profile_id).first()
    if not row:
        raise HTTPException(404, "Profile not found")

    name = (name or "").strip()
    if not name:
        raise HTTPException(400, "Name is required")

    if photo and (photo.filename or "").strip():
        content_type = (photo.content_type or "").lower()
        if not content_type.startswith("image/"):
            raise HTTPException(400, "Photo must be an image")
        ext = Path(photo.filename).suffix.lower() if photo.filename else ".jpg"
        if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            ext = ".jpg"
        old_gcs_path = url_to_gcs_path(row.photo_url)
        if old_gcs_path:
            delete_file(old_gcs_path)
        filename = generate_filename("team", ext)
        destination = f"team/{filename}"
        row.photo_url = upload_file(photo.file, destination, content_type)

    row.name = name
    row.role_title = (role_title or "").strip()
    row.bio = (bio or "").strip() or None
    row.display_order = int(display_order)
    row.is_active = bool(is_active)
    db.commit()
    db.refresh(row)
    return _profile_row(row)


@router.delete("/team-profiles/{profile_id}")
def delete_team_profile(
    profile_id: int,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == profile_id).first()
    if not row:
        raise HTTPException(404, "Profile not found")
    old_gcs_path = url_to_gcs_path(row.photo_url)
    if old_gcs_path:
        delete_file(old_gcs_path)
    db.delete(row)
    db.commit()
    return {"success": True, "profile_id": profile_id}


# ── Portfolio Configuration ───────────────────────────────────────────────────

@router.get("/public/portfolio")
def public_get_portfolio(db: Session = Depends(get_db)):
    raw = _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, PORTFOLIO_PARAM_KEY)
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}


@router.get("/portfolio")
def get_portfolio(
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    raw = _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, PORTFOLIO_PARAM_KEY)
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}


@router.put("/portfolio")
def update_portfolio(
    payload: dict,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    _set_param_value(
        db, PLATFORM_SETTINGS_SHOP_ID, PORTFOLIO_PARAM_KEY, json.dumps(payload, ensure_ascii=False),
    )
    db.commit()
    return {"success": True}


@router.post("/portfolio/photo")
def upload_portfolio_photo(
    photo: UploadFile = File(...),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    ext = Path(photo.filename).suffix.lower() if photo.filename else ".jpg"
    filename = generate_filename("portfolio_photo", ext)
    destination = f"platform/{filename}"
    photo_url = upload_file(photo.file, destination, photo.content_type or "image/jpeg")
    return {"success": True, "photo_url": photo_url}


# ── Portfolios (multi) ──────────────────────────────────────────────────────
# Each portfolio is a standalone slug + JSON config, optionally linked to a
# team profile so its name becomes clickable on the public About page.

LEGACY_PORTFOLIO_SLUG = "sathish_kumar_lakshman"


def _slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (text or "").strip().lower()).strip("-")
    return s or "portfolio"


def _unique_portfolio_slug(db: Session, base_slug: str, exclude_id: int | None = None) -> str:
    slug = base_slug
    i = 2
    while True:
        q = db.query(PlatformPortfolio).filter(PlatformPortfolio.slug == slug)
        if exclude_id is not None:
            q = q.filter(PlatformPortfolio.portfolio_id != exclude_id)
        if not q.first():
            return slug
        slug = f"{base_slug}-{i}"
        i += 1


def _portfolio_config_dict(row: PlatformPortfolio) -> dict:
    try:
        return json.loads(row.config_json) if row.config_json else {}
    except (json.JSONDecodeError, TypeError):
        return {}


def _portfolio_summary(row: PlatformPortfolio, profile_name: str | None = None) -> dict:
    return {
        "portfolio_id": row.portfolio_id,
        "slug": row.slug,
        "profile_id": row.profile_id,
        "profile_name": profile_name,
        "is_active": row.is_active,
        "created_at": str(row.created_at) if row.created_at else None,
        "updated_at": str(row.updated_at) if row.updated_at else None,
    }


def _get_or_migrate_legacy_portfolio(db: Session, slug: str) -> PlatformPortfolio | None:
    """Look up a portfolio row by slug; for the legacy default slug, lazily
    migrate the old single-config value (portfolio_config system param) into
    the new table on first access so existing content isn't lost."""
    row = db.query(PlatformPortfolio).filter(PlatformPortfolio.slug == slug).first()
    if row:
        return row
    if slug != LEGACY_PORTFOLIO_SLUG:
        return None
    raw = _get_param_value(db, PLATFORM_SETTINGS_SHOP_ID, PORTFOLIO_PARAM_KEY)
    row = PlatformPortfolio(slug=LEGACY_PORTFOLIO_SLUG, config_json=raw or "{}", is_active=True)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/public/portfolios/{slug}")
def public_get_portfolio_by_slug(slug: str, db: Session = Depends(get_db)):
    row = _get_or_migrate_legacy_portfolio(db, slug)
    if not row or not row.is_active:
        raise HTTPException(404, "Portfolio not found")
    return _portfolio_config_dict(row)


@router.get("/portfolios")
def list_portfolios(
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    _get_or_migrate_legacy_portfolio(db, LEGACY_PORTFOLIO_SLUG)
    rows = db.query(PlatformPortfolio).order_by(PlatformPortfolio.created_at.asc()).all()
    profile_ids = [r.profile_id for r in rows if r.profile_id is not None]
    names = {}
    if profile_ids:
        for p in db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id.in_(profile_ids)).all():
            names[p.profile_id] = p.name
    return [_portfolio_summary(r, names.get(r.profile_id)) for r in rows]


@router.get("/portfolios/{slug}")
def get_portfolio_by_slug(
    slug: str,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = _get_or_migrate_legacy_portfolio(db, slug)
    if not row:
        raise HTTPException(404, "Portfolio not found")
    return _portfolio_config_dict(row)


class PortfolioCreateRequest(BaseModel):
    slug: str | None = None
    name: str | None = None
    profile_id: int | None = None


@router.post("/portfolios")
def create_portfolio(
    payload: PortfolioCreateRequest,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    base = payload.slug or payload.name
    if payload.profile_id and not base:
        profile = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == payload.profile_id).first()
        if profile:
            base = profile.name
    base_slug = _slugify(base or "portfolio")
    slug = _unique_portfolio_slug(db, base_slug)

    config = {}
    if payload.profile_id:
        profile = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == payload.profile_id).first()
        if not profile:
            raise HTTPException(404, "Team profile not found")
        config = {
            "hero_name": profile.name,
            "hero_title_line1": profile.role_title or "",
            "hero_title_line2": "",
            "photo_url": profile.photo_url or "",
        }

    row = PlatformPortfolio(
        slug=slug,
        profile_id=payload.profile_id,
        config_json=json.dumps(config, ensure_ascii=False),
        is_active=True,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    profile_name = None
    if row.profile_id:
        p = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == row.profile_id).first()
        profile_name = p.name if p else None
    return _portfolio_summary(row, profile_name)


@router.put("/portfolios/{slug}")
def upsert_portfolio_content(
    slug: str,
    payload: dict,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = _get_or_migrate_legacy_portfolio(db, slug)
    if not row:
        row = PlatformPortfolio(slug=slug, config_json="{}", is_active=True)
        db.add(row)
    row.config_json = json.dumps(payload, ensure_ascii=False)
    db.commit()
    return {"success": True}


class PortfolioMetaUpdate(BaseModel):
    profile_id: int | None = None
    is_active: bool | None = None


@router.patch("/portfolios/{slug}")
def update_portfolio_meta(
    slug: str,
    payload: PortfolioMetaUpdate,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = _get_or_migrate_legacy_portfolio(db, slug)
    if not row:
        raise HTTPException(404, "Portfolio not found")
    fields = payload.dict(exclude_unset=True)
    if "profile_id" in fields:
        pid = fields["profile_id"]
        if pid:
            profile = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == pid).first()
            if not profile:
                raise HTTPException(404, "Team profile not found")
            row.profile_id = pid
        else:
            row.profile_id = None
    if "is_active" in fields:
        row.is_active = fields["is_active"]
    db.commit()
    db.refresh(row)
    profile_name = None
    if row.profile_id:
        p = db.query(PlatformTeamProfile).filter(PlatformTeamProfile.profile_id == row.profile_id).first()
        profile_name = p.name if p else None
    return _portfolio_summary(row, profile_name)


@router.delete("/portfolios/{slug}")
def delete_portfolio(
    slug: str,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    row = db.query(PlatformPortfolio).filter(PlatformPortfolio.slug == slug).first()
    if not row:
        raise HTTPException(404, "Portfolio not found")
    db.delete(row)
    db.commit()
    return {"success": True, "slug": slug}


@router.post("/portfolios/{slug}/photo")
def upload_portfolio_photo_by_slug(
    slug: str,
    photo: UploadFile = File(...),
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    ext = Path(photo.filename).suffix.lower() if photo.filename else ".jpg"
    filename = generate_filename("portfolio", ext)
    destination = f"platform/{filename}"
    photo_url = upload_file(photo.file, destination, photo.content_type or "image/jpeg")
    return {"success": True, "photo_url": photo_url}


@router.post("/shops/{shop_id}/backup")
def export_shop_backup(
    shop_id: int,
    payload: dict | None = None,
    db: Session = Depends(get_db),
    owner=Depends(PlatformOwnerOnly),
):
    from app.models.invoice import Invoice
    from app.models.invoice_details import InvoiceDetail
    from app.models.items import Item
    from app.models.category import Category
    from app.models.branch import Branch
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    shop = db.query(ShopDetails).filter(ShopDetails.shop_id == shop_id).first()
    if not shop:
        raise HTTPException(404, "Shop not found")

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B3C8C", end_color="0B3C8C", fill_type="solid")

    def add_sheet(name, columns, rows):
        ws = wb.create_sheet(title=name[:31])
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([str(v) if v is not None else "" for v in row])
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # Shop Details
    add_sheet("Shop", ["Shop ID", "Shop Name", "Owner", "Mobile", "Email", "GST No", "Plan", "Billing Type"],
              [[shop.shop_id, shop.shop_name, shop.owner_name, shop.mobile, shop.mailid, shop.gst_number, shop.plan, shop.billing_type]])

    # Branches
    branches = db.query(Branch).filter(Branch.shop_id == shop_id).all()
    add_sheet("Branches", ["ID", "Name", "Address", "City", "State", "Pincode", "Type", "Status"],
              [[b.branch_id, b.branch_name, getattr(b, "address_line1", ""), getattr(b, "city", ""),
                getattr(b, "state", ""), getattr(b, "pincode", ""), b.type, b.status] for b in branches])

    # Categories
    cats = db.query(Category).filter(Category.shop_id == shop_id).all()
    add_sheet("Categories", ["ID", "Name", "Status"],
              [[c.category_id, c.category_name, c.category_status] for c in cats])

    # Items
    items = db.query(Item).filter(Item.shop_id == shop_id).all()
    add_sheet("Items", ["ID", "Name", "Category ID", "Price", "Buy Price", "MRP", "HSN", "GST Rate", "Unit", "Status"],
              [[i.item_id, i.item_name, i.category_id, float(i.price or 0), float(i.buy_price or 0),
                float(getattr(i, "mrp_price", 0) or 0), i.hsn_code, float(i.gst_rate or 0),
                i.unit, i.item_status] for i in items])

    # Invoices
    invoices = db.query(Invoice).filter(Invoice.shop_id == shop_id).order_by(Invoice.created_time.desc()).limit(5000).all()
    add_sheet("Invoices", ["ID", "Invoice No", "Branch ID", "Customer", "Mobile", "Total", "Discount", "Tax", "Payment", "Date"],
              [[inv.invoice_id, inv.invoice_number, inv.branch_id, inv.customer_name, inv.mobile,
                float(inv.total_amount or 0), float(inv.discounted_amt or 0), float(inv.tax_amt or 0),
                inv.payment_mode, str(inv.created_time or "")] for inv in invoices])

    # Invoice Details
    inv_ids = [inv.invoice_id for inv in invoices]
    if inv_ids:
        details = db.query(InvoiceDetail).filter(InvoiceDetail.invoice_id.in_(inv_ids)).all()
        add_sheet("Invoice Items", ["Invoice ID", "Item ID", "Item Name", "Qty", "Price", "Amount"],
                  [[d.invoice_id, d.item_id, getattr(d, "item_name", ""), float(getattr(d, "quantity", 0) or 0),
                    float(getattr(d, "price", 0) or 0), float(getattr(d, "amount", 0) or 0)] for d in details])

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.getvalue()

    # Email if requested
    send_email = (payload or {}).get("send_email", False)
    if send_email and shop.mailid and _can_send_mail():
        from email.mime.base import MIMEBase
        from email import encoders

        msg = EmailMessage()
        msg["Subject"] = f"Shop Data Backup — {shop.shop_name}"
        msg["From"] = SENDER_EMAIL
        msg["To"] = shop.mailid
        msg.set_content(f"Hi,\n\nPlease find attached the data backup for {shop.shop_name}.\n\nRegards,\nHaappii Billing")
        msg.add_attachment(excel_bytes, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          filename=f"backup_{shop.shop_name}_{date.today().isoformat()}.xlsx")
        try:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
        except Exception:
            pass

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="backup_{shop.shop_name}_{date.today().isoformat()}.xlsx"'},
    )
