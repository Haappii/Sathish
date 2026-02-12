from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
import os

from app.db import get_db
from app.models.items import Item
from app.models.stock import Inventory

from app.services.inventory_service import (
    is_inventory_enabled,
    ensure_stock_row,
    adjust_stock,
)

from app.utils.auth_user import get_current_user
from app.services.audit_service import log_action

router = APIRouter(prefix="/inventory", tags=["Inventory Bulk Upload"])


def resolve_branch(branch_id_param, user):
    if str(user.role_name).lower() == "admin":
        return int(branch_id_param or user.branch_id)
    return int(user.branch_id)


@router.get("/bulk-template")
def download_bulk_template():
    os.makedirs("uploads", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Upload"

    ws.append(["Item Name", "Quantity", "A / R"])
    ws.append(["Example Item 1", "10", "A"])
    ws.append(["Example Item 2", "5", "R"])

    path = "uploads/inventory_bulk_template.xlsx"
    wb.save(path)

    return FileResponse(
        path,
        filename="inventory_bulk_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/bulk-upload")
async def bulk_upload_stock(
    file: UploadFile = File(...),
    branch_id: int | None = Query(None),
    db: Session = Depends(get_db),
    user = Depends(get_current_user)
):

    if not is_inventory_enabled(db, user.shop_id):
        raise HTTPException(400, "Inventory mode disabled")

    branch = resolve_branch(branch_id, user)

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")

    os.makedirs("uploads", exist_ok=True)
    temp_path = f"uploads/{file.filename}"

    with open(temp_path, "wb") as f:
        f.write(await file.read())

    wb = load_workbook(temp_path)
    sheet = wb.active

    processed = 0
    failed = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_name, qty, mode = row

        if not item_name or not qty or not mode:
            failed.append({"item": item_name, "reason": "Missing fields"})
            continue

        item = (
            db.query(Item)
            .filter(
                Item.item_name == str(item_name).strip(),
                Item.shop_id == user.shop_id
            )
            .first()
        )

        if not item:
            failed.append({"item": item_name, "reason": "Item not found"})
            continue

        ensure_stock_row(db, user.shop_id, item.item_id, branch)

        qty = int(qty)
        mode = str(mode).upper()

        if mode == "A":
            adjust_stock(db, user.shop_id, item.item_id, branch, qty, "ADD")

        elif mode == "R":
            ok = adjust_stock(db, user.shop_id, item.item_id, branch, qty, "REMOVE")
            if not ok:
                failed.append({"item": item_name, "reason": "Insufficient stock"})
                continue

        else:
            failed.append({"item": item_name, "reason": "Invalid Mode (Use A / R)"})
            continue

        processed += 1

    db.commit()
    os.remove(temp_path)

    log_action(
        db,
        shop_id=user.shop_id,
        module="Inventory",
        action="BULK_UPLOAD",
        record_id=f"branch:{branch}",
        new={
            "branch_id": branch,
            "processed": processed,
            "failed_count": len(failed),
            "filename": file.filename,
        },
        user_id=user.user_id,
    )

    return {
        "success": True,
        "branch_id": branch,
        "processed": processed,
        "failed": failed
    }
